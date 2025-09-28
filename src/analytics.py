# src/analytics.py
import os
from datetime import datetime

import pandas as pd
import matplotlib.pyplot as plt
from sqlalchemy import create_engine, text

from src.config import DB_URL



# ---------- общие настройки ----------
os.makedirs("charts", exist_ok=True)
os.makedirs("exports", exist_ok=True)

# единый лаконичный стиль для всех графиков (требование: consistent style)
plt.rcParams.update({
    "figure.figsize": (9, 5),
    "axes.grid": True,
    "axes.titlesize": 12,
    "axes.labelsize": 10
})

engine = create_engine(DB_URL, future=True)

def df_from_sql(sql: str, params: dict | None = None) -> pd.DataFrame:
    with engine.connect() as conn:
        return pd.read_sql(text(sql), conn, params=params or {})

def save_and_report(df: pd.DataFrame, fig, fname: str, about: str, kind: str):
    out = os.path.join("charts", fname)
    fig.savefig(out, bbox_inches="tight")
    plt.close(fig)
    print(f"[OK] rows={len(df)} | type={kind} | file={out} | about={about}")


# ============= 6 ГРАФИКОВ (каждый с ≥2 JOIN) =============

def chart_pie_revenue_by_category():
    sql = """
    SELECT p.category AS category_name,
           SUM(oi.quantity * oi.price_at_purchase) AS revenue
    FROM order_items oi
    JOIN products p ON p.product_id = oi.product_id
    JOIN orders o   ON o.order_id   = oi.order_id
    GROUP BY p.category
    ORDER BY revenue DESC;
    """
    df = df_from_sql(sql)
    fig = plt.figure()
    ax = fig.gca()
    ax.pie(df["revenue"], labels=df["category_name"], autopct="%1.1f%%")
    ax.set_title("Revenue share by category")
    save_and_report(df, fig, "pie_revenue_by_category.png",
                    "Revenue distribution across categories",
                    "pie")



def chart_bar_top_products():
    sql = """
    SELECT p.product_name,
           SUM(oi.quantity * oi.price_at_purchase) AS revenue
    FROM order_items oi
    JOIN products p ON p.product_id = oi.product_id
    JOIN orders o   ON o.order_id   = oi.order_id
    GROUP BY p.product_name
    ORDER BY revenue DESC
    LIMIT 10;
    """
    df = df_from_sql(sql)
    fig = plt.figure()
    ax = fig.gca()
    ax.bar(df["product_name"], df["revenue"])
    ax.set_title("Top-10 products by revenue")
    ax.set_xlabel("Product")
    ax.set_ylabel("Revenue")
    plt.xticks(rotation=30, ha="right")
    save_and_report(df, fig, "bar_top_products.png",
                    "Top-10 products by revenue",
                    "bar")



def chart_barh_aov_by_address():
    sql = """
    WITH order_totals AS (
      SELECT o.order_id, SUM(oi.quantity*oi.price_at_purchase) AS order_total
      FROM orders o
      JOIN order_items oi ON oi.order_id = o.order_id
      GROUP BY o.order_id
    )
    SELECT c.address AS address_name,
           AVG(ot.order_total) AS avg_order_value
    FROM orders o
    JOIN customers c    ON c.customer_id = o.customer_id
    JOIN order_totals ot ON ot.order_id  = o.order_id
    GROUP BY c.address
    ORDER BY avg_order_value DESC
    LIMIT 10;
    """
    df = df_from_sql(sql)
    fig = plt.figure()
    ax = fig.gca()
    ax.barh(df["address_name"], df["avg_order_value"])
    ax.set_title("Average order value by address")
    ax.set_xlabel("Average order value")
    save_and_report(df, fig, "barh_aov_by_address.png",
                    "Average order value by address",
                    "barh")



def chart_line_monthly_revenue_by_category():
    sql = """
    SELECT DATE_TRUNC('month', o.order_date)::date AS month,
           p.category AS category_name,
           SUM(oi.quantity*oi.price_at_purchase) AS revenue
    FROM orders o
    JOIN order_items oi ON oi.order_id   = o.order_id
    JOIN products p     ON p.product_id  = oi.product_id
    GROUP BY month, p.category
    ORDER BY month, p.category;
    """
    df = df_from_sql(sql)
    pivot = df.pivot(index="month", columns="category_name", values="revenue").fillna(0)
    fig = plt.figure()
    ax = fig.gca()
    pivot.plot(ax=ax)
    ax.set_title("Monthly revenue by category")
    ax.set_xlabel("Month")
    ax.set_ylabel("Revenue")
    ax.legend(title="Category")
    save_and_report(df, fig, "line_monthly_rev_by_category.png",
                    "Monthly revenue by category (trend)",
                    "line")


def chart_hist_order_totals():
    sql = """
    SELECT o.order_id,
           SUM(oi.quantity*oi.price_at_purchase) AS order_total
    FROM orders o
    JOIN order_items oi ON oi.order_id = o.order_id
    JOIN customers c    ON c.customer_id = o.customer_id
    GROUP BY o.order_id;
    """
    df = df_from_sql(sql)
    fig = plt.figure()
    ax = fig.gca()
    ax.hist(df["order_total"], bins=20)
    ax.set_title("Distribution of order totals")
    ax.set_xlabel("Order total")
    ax.set_ylabel("Count of orders")
    save_and_report(df, fig, "hist_order_totals.png",
                    "Distribution of order totals",
                    "hist")



def chart_scatter_price_vs_qty():
    sql = """
    SELECT p.product_name,
           p.price AS product_price,
           SUM(oi.quantity) AS qty_sold
    FROM order_items oi
    JOIN products p ON p.product_id = oi.product_id
    JOIN orders o   ON o.order_id   = oi.order_id
    GROUP BY p.product_name, p.price
    ORDER BY qty_sold DESC;
    """
    df = df_from_sql(sql)
    fig = plt.figure()
    ax = fig.gca()
    ax.scatter(df["product_price"], df["qty_sold"])
    ax.set_title("Price vs quantity sold")
    ax.set_xlabel("Product price")
    ax.set_ylabel("Quantity sold")
    save_and_report(df, fig, "scatter_price_vs_qty.png",
                    "Price vs sales volume (by product)",
                    "scatter")



# ============= Plotly: Time Slider (анимация по времени) =============
# Требование: интерактивный график с ползунком времени через animation_frame,
# сохранять не нужно, просто показать интерактивность. 
# (можно генерировать дату из order_date по месяцам) 
# См. подсказку задания с animation_frame. 
import plotly.express as px  # noqa: E402

def plotly_time_slider_sales_by_month_and_category(show=True):
    sql = """
    SELECT DATE_TRUNC('month', o.order_date)::date AS month,
           p.category AS category_name,
           SUM(oi.quantity*oi.price_at_purchase) AS revenue
    FROM orders o
    JOIN order_items oi ON oi.order_id   = o.order_id
    JOIN products p     ON p.product_id  = oi.product_id
    GROUP BY month, p.category
    ORDER BY month, p.category;
    """
    df = df_from_sql(sql)
    fig = px.bar(
        df, x="category_name", y="revenue",
        animation_frame=df["month"].astype(str),
        title="Monthly revenue by category (time slider)"
    )
    if show:
        fig.show()
    return fig



# ============= Экспорт в Excel с форматированием (openpyxl) =============
from openpyxl import load_workbook  # noqa: E402
from openpyxl.formatting.rule import ColorScaleRule  # noqa: E402

def export_to_excel(dfs: dict[str, pd.DataFrame], filename: str):
    """
    Пишем несколько DataFrame на разные листы, затем применяем:
    - freeze panes (B2),
    - автофильтры,
    - градиент по числовым колонкам (ColorScaleRule),
    - выводим консольный отчёт.
    Сохраняем в /exports/.
    """
    path = os.path.join("exports", filename)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        total_rows = 0
        for sheet, df in dfs.items():
            df.to_excel(writer, sheet_name=sheet, index=False)
            total_rows += len(df)
    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "B2"  # закрепляем заголовки
        ws.auto_filter.ref = ws.dimensions  # включаем фильтры на все колонки

        # Применим цветовую шкалу к диапазону со 2-й строки, начиная с C-колонки
        if ws.max_row >= 2 and ws.max_column >= 3:
            last_col_letter = ws.cell(row=1, column=ws.max_column).column_letter
            rng = f"C2:{last_col_letter}{ws.max_row}"
            rule = ColorScaleRule(
                start_type="min", start_color="FFAA0000",
                mid_type="percentile", mid_value=50, mid_color="FFFFFF00",
                end_type="max", end_color="FF00AA00"
            )
            ws.conditional_formatting.add(rng, rule)
    wb.save(path)
    print(f"[OK] Created file {filename}, {len(dfs)} sheets, {total_rows} rows")


# ============= Демо для защиты: вставка строки и пересчёт =============
def demo_insert_one_order():
    """
    1) Создаём новый заказ (orders)
    2) Добавляем одну позицию (order_items) с любой ценой/количеством.
    !! ВАЖНО: подставь существующие customer_id и product_id !!
    """
    with engine.begin() as conn:
        new_order_id = conn.execute(text("""
            INSERT INTO orders (customer_id, order_date, status)
            VALUES (:cid, NOW()::timestamp, 'completed')
            RETURNING order_id
        """), {"cid": 1}).scalar_one()

        conn.execute(text("""
            INSERT INTO order_items (order_id, product_id, quantity, unit_price)
            VALUES (:oid, :pid, :q, :price)
        """), {"oid": new_order_id, "pid": 1, "q": 1, "price": 100})

    print(f"[OK] inserted demo order_id={new_order_id}")


# ============= Точка входа =============
if __name__ == "__main__":
    # 6 графиков:
    chart_pie_revenue_by_category()
    chart_bar_top_products()
    chart_barh_aov_by_address()
    chart_line_monthly_revenue_by_category()
    chart_hist_order_totals()
    chart_scatter_price_vs_qty()


    # Пример экспорта:
    df_orders = df_from_sql("SELECT * FROM orders LIMIT 1000;")
    df_items  = df_from_sql("""
        SELECT oi.*, p.product_name
        FROM order_items oi
        JOIN products p ON p.product_id = oi.product_id
        LIMIT 1000;
    """)
    export_to_excel({"orders": df_orders, "items": df_items}, "nomadshop_sample.xlsx")

    # Plotly time slider — вызывать вручную при защите / отладки:
    # plotly_time_slider_sales_by_month_and_category(show=True)

    # Демо вставки строки — вызывать на защите между двумя генерациями графика:
    # demo_insert_one_order()
# --- DIAG: посмотреть схему БД без psql ---


# --- DIAG: посмотреть схему БД без psql ---
from sqlalchemy import inspect

def print_schema():
    insp = inspect(engine)
    tables = insp.get_table_names(schema="public")
    print("\n== TABLES ==")
    for t in tables:
        print(" -", t)
    print("\n== COLUMNS by table ==")
    for t in tables:
        cols = insp.get_columns(t, schema="public")
        print(f"\n[{t}]")
        for c in cols:
            print(f"  {c['name']:20} {c.get('type')}")

def list_tables_sql():
    sql = """
    SELECT table_name
    FROM information_schema.tables
    WHERE table_schema = 'public'
    ORDER BY 1;
    """
    df = df_from_sql(sql)
    print(df.to_string(index=False))

def describe_table(name: str):
    sql = """
    SELECT column_name, data_type
    FROM information_schema.columns
    WHERE table_schema='public' AND table_name = :t
    ORDER BY ordinal_position;
    """
    df = df_from_sql(sql, {"t": name})
    print(f"\n[{name}]")
    print(df.to_string(index=False))
